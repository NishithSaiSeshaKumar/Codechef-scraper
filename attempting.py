import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side
from concurrent.futures import ThreadPoolExecutor
import requests
from bs4 import BeautifulSoup

def analyze_attendance_and_apply_colors(input_file, output_file, contest_no, max_workers=10):
    # Read the Excel file
    DF = pd.read_excel(input_file)
    DF['YearSection'] = DF['Year'].astype(str) + DF['Section ']

    # Create a new DataFrame for analysis
    DF2 = pd.DataFrame()

    # Calculate attempted, not attempted, and attendance percentage
    count_not_attempted = DF[~DF['CONTEST 127'].str.contains("not", case=False, na=False)].groupby('YearSection').count() / DF.groupby('YearSection').count()

    # Add data to DF2
    DF2['YearSection'] = DF['YearSection'].unique()
    DF2['Attempted'] = DF[~DF['CONTEST 127'].str.contains("not", case=False, na=False)].groupby('YearSection').count()['CONTEST 127']
    DF2['NotAttempted'] = DF.groupby('YearSection').count()['CONTEST 127'] - DF2['Attempted']
    DF2['Attendace %'] = count_not_attempted['CONTEST 127'] * 100
    DF2['Total'] = DF.groupby('YearSection').count()['CONTEST 127']
    
    # Save the analysis DataFrame to a new sheet in the same Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
        DF2.to_excel(writer, sheet_name=f'Analysis_{contest_no}', index=False)


    # Apply colors and styling to the new sheet
    wb = openpyxl.load_workbook(output_file)
    sheet = wb[f'Analysis_{contest_no}']

    colors = ['#FFFFFACD', '#FFB0E0E6', '#FF90EE90', '#FFE4E1']
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    font_size = 20

    # Define function to apply colors and styling to rows
    def apply_colors_and_styling(row):
        nonlocal colors
        nonlocal border_style
        nonlocal font_size
        color = colors[row + 1 % len(colors)]  # Start with index 1 to avoid header row
        for cell in row:
            cell.fill = PatternFill(start_color=color[1:], end_color=color[1:], fill_type='solid')
            cell.border = border_style
            cell.font = Font(size=font_size)
            if row[0].row == 1:  # Header row
                cell.font = Font(size=font_size, bold=True)

    # Apply colors and styling to each row
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        executor.map(apply_colors_and_styling, sheet.iter_rows(min_row=2, max_row=sheet.max_row))

    # Save the modified Excel file
    wb.save(output_file)

def attempt(users, contest_no, max_workers=10):
    if not users or not all(users):
        return [], []

    def attempt_single(user):
        if not user:
            return ""
        url = 'https://www.codechef.com/users/' + user.replace(" ", "%20")
        try:
            contest_response = requests.head(url)
            if contest_response.status_code == 200:
                with requests.Session() as session:
                    r = session.get(url)
                    r.raise_for_status()
                    soup = BeautifulSoup(r.content, 'html.parser')
                    rating = soup.find('a', class_='rating')
                    contest = soup.find('div', class_='contest-name')
                    rate = rating.text.split()[0].split("?")[0] if hasattr(rating, 'text') else " "
                    if contest and hasattr(contest, 'text'):
                        if contest_no in contest.text:
                            return f"Attempted ({rate})"
                        else:
                            return f"Not Attempted ({rate})"
                    else:
                        return ""
            else:
                return ""
        except requests.RequestException as e:
            print(f"Error accessing URL: {url} - {e}")
            return f"Error {user}"

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        results = list(executor.map(attempt_single, users))

    return results, []
